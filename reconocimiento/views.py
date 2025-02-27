from django.shortcuts import render, redirect
from PIL import Image
import pytesseract
import datetime
import io
from .models import Tipo, Trabajador, Lugar, Empresa, Visita
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.db.models import Count
import cv2
import numpy as np
from django.core.paginator import Paginator

pytesseract.pytesseract.tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'


#Variables globales
nombre, apellido_paterno, apellido_materno, RUT = None, None, None, None
hora_actual, tipo_visitante_id, observacion, fecha_actual = None, None, None, None
lugar_ID, fecha_salida, empresa_ID, trabajador_ID, visitas_filtradas = None, None, None, None, None
visitas_filtradas=Visita.objects.all()

def index(request):
    return render(request, 'index.html')


def procesar_imagen(request):
    
    global nombre, apellido_paterno, apellido_materno, RUT, fecha_actual, hora_actual, tipo_visitante_id, observacion, lugar_ID, fecha_salida, empresa_ID, trabajador_ID
    fecha_salida=None

    if request.method != 'POST':
        return render(request, 'index.html')

    img_data = request.POST.get('img_data')
    if not img_data:
        return render(request, 'index.html', {'error_message': "No se recibió una imagen válida."})
    
    img_data = img_data.replace("data:image/png;base64,", "")
    image = decode_image(img_data)
    if not image:
        return render(request, 'index.html', {'error_message': "Error al procesar la imagen."})
    
    texto = extraer_texto_de_imagen(image)
    if not texto:
        return render(request, 'index.html', {'error_message': "Verifica si la imagen es legible o el archivo está en el formato correcto."})
    
    lista_filtrada = obtener_lista_filtrada(texto)
    if len(lista_filtrada) <= 1:
        return render(request, 'index.html', {'error_message': "No se reconoció la imagen."})
    
    RUT, nombres = obtener_rut_y_nombres(lista_filtrada)
    if len(nombres) < 3: #no hay separacion entre nombre y apellidos
        return render(request, 'index.html', {'error_message': "No se reconoció la imagen."})
    
    try:
        if int(RUT) <= 3000000:
            return render(request, 'index.html', {'error_message': "No se reconoció la imagen."})
    except ValueError:
        return render(request, 'index.html', {'error_message': "No se reconoció la imagen."})
    
    nombre, apellido_paterno, apellido_materno = nombres[2].title(), nombres[0].title(), nombres[1].title()
    fecha_actual = datetime.datetime.now()
    hora_actual = fecha_actual.time()
    
    ultima_visita = Visita.objects.filter(visita_rut=RUT).order_by('-visita_fecha_llegada').first()
    if ultima_visita and ultima_visita.visita_activo:
        ultima_visita.visita_fecha_salida = fecha_actual
        ultima_visita.visita_activo = False
        ultima_visita.save()
        return redirect(registro_salida)
    
    return redirect(datos)


def obtener_lista_filtrada(texto): #eliminar todo el texto que no se necesita

    lineas = texto.split('\n')
    lista_filtrada = []
    for i in range(len(lineas)):
        if lineas[i].startswith("INCHL"): #todos los carnets chilenos inician los datos con INCHL
            lista_filtrada.append(lineas[i+1])
            if i + 2 < len(lineas):
                if lineas[i+2] == "":
                    if i + 3 < len(lineas):
                        lista_filtrada.append(lineas[i+3])
                else:
                    lista_filtrada.append(lineas[i+2])
    return lista_filtrada


def obtener_rut_y_nombres(lista_filtrada):
    if len(lista_filtrada) < 2: #si la imagen no cumple con las dimensiones
        RUT=[]
        nombres=[]
        return RUT, nombres

    rut_separado = lista_filtrada[0].split("CHL") #rut inicia despues del CHL
    if len(rut_separado) < 2: #Si no lo separa
        RUT=[]
        nombres=[]
        return RUT, nombres

    RUT_con_espacios = rut_separado[1].split("<")[0] #separa la linea de rut

    RUT = RUT_con_espacios.replace(" ", "") #si detecta espacios

    nombres_separados = lista_filtrada[1].split("<") #segunda linea separa apellidos y nombres

    nombres = [nombre for nombre in nombres_separados if nombre != ""] #si es distinto de nulo lo agrega a la lista

    for nombre in nombres: #verifica si hay espacios entre las letras de un nombre
        if " " in nombre:
            RUT=[]
            nombres=[]
            return RUT, nombres
    if len(nombres) > 2:    
        if not (nombres[0].isalpha() and nombres[1].isalpha() and nombres[2].isalpha()) or nombres[0].startswith(' ') or nombres[1].startswith(' ') or nombres[2].startswith(' '): #Los nombres o apellidos no pueden comenzar con caracteres no alfabéticos o espacios
                RUT=[]
                nombres=[]
                return RUT, nombres                        
    return RUT, nombres


def extraer_texto_de_imagen(imagen):
    try:
        image_np = np.array(imagen)
        # Convertir la imagen a escala de grises
        gray = cv2.cvtColor(image_np, cv2.COLOR_BGR2GRAY)

        # Aplicar el filtro blackhat
        kernel = np.ones((100,100),np.uint8)
        blackhat = cv2.morphologyEx(gray, cv2.MORPH_BLACKHAT, kernel)

        # Realizar el proceso de extracción de texto sobre la imagen con el filtro aplicado
        texto = pytesseract.image_to_string(blackhat)
        
        return texto
    except Exception as e:
        print(f"Error al extraer texto: {str(e)}")
        return None


def decode_image(img_data):
    try:
        import base64
        img_bytes = base64.b64decode(img_data)
        image = Image.open(io.BytesIO(img_bytes))
        return image
    except Exception as e:
        print(f"Error al decodificar imagen: {str(e)}")
        return None


def finalizar_visita_manual(request):
    global nombre,apellido_paterno,apellido_materno,RUT,hora_actual,fecha_actual

    fecha_actual = datetime.datetime.now()
    hora_actual = fecha_actual.time()  

    if request.method == 'POST':
        visita_id = request.POST.get('visita_id')
        nombre = request.POST.get('nombre')
        apellido_paterno = request.POST.get('apellido_paterno')
        apellido_materno = request.POST.get('apellido_materno')
        RUT = request.POST.get('RUT')
        if visita_id:
            visita = Visita.objects.get(pk=visita_id)
            if visita.visita_activo==True:
                visita.visita_fecha_salida = datetime.datetime.now()
                visita.visita_activo=False
                visita.save()
                return redirect(registro_salida)
        else:
            return redirect(index)


def registro_salida(request):
    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'hora_actual': hora_actual,
        'fecha_actual': fecha_actual,
        }                           
    return render(request, 'registro_salida.html', context)


def datos(request): 
    global observacion, empresa_ID, trabajador_ID

    # Reiniciar valores de variables
    empresa_ID,observacion, trabajador_ID = None, None, None

    # Consultar la base de datos para obtener los tipos de visitantes
    tipos_de_visitantes = Tipo.objects.all()
    context = {
        'tipos_de_visitantes': tipos_de_visitantes,
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'fecha_actual': fecha_actual,
        'hora_actual': hora_actual,
        }                           
    return render(request, 'datos.html', context)


def apoderado(request):
    global tipo_visitante_id
    trabajadores = Trabajador.objects.all()
    tipo_visitante_id=1

    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'trabajadores': trabajadores,
    }
    return render(request, 'colaborador.html', context)


def otras_visitas(request):
    global tipo_visitante_id
    trabajadores = Trabajador.objects.all()
    tipo_visitante_id=4

    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'trabajadores': trabajadores,
    }
    return render(request, 'colaborador.html', context)


def entrevista(request):
    global tipo_visitante_id
    trabajadores = Trabajador.objects.all()
    tipo_visitante_id=3

    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'trabajadores': trabajadores,
        
    }
    return render(request, 'colaborador.html', context)


def contratista(request):
    global tipo_visitante_id, observacion
    nombre_empresas = Empresa.objects.all()
    tipo_visitante_id=2

    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'fecha_actual': fecha_actual,
        'nombre_empresas': nombre_empresas,
    }
    return render(request, 'contratista.html', context)


def procesar_post_contratista(request):
    global observacion, empresa_ID
    if request.method == 'POST':
        empresa_ID = request.POST.get('idEmpresa', None)
        observacion = request.POST.get('observacionHidden', None)
    return redirect(ubicacion)


def agregar_empresa(request):
    if request.method == "POST":
        nombre_empresa = request.POST.get('empresa')
        # Verifica que el nombre de la empresa no esté vacío o contenga solo espacios en blanco
        if nombre_empresa and not nombre_empresa.isspace():
            # Crea una nueva instancia de Empresa
            nueva_empresa = Empresa(empresa_nombre=nombre_empresa)
            nueva_empresa.save()
            return redirect(contratista)  # Redirige a la página actual
    return redirect(contratista)  # Si no se proporciona un nombre, redirige a la página actual


def procesar_post_colaborador(request):
    global trabajador_ID, observacion
    if request.method == 'POST':
        trabajador_ID = request.POST.get('idColaborador', None)
        observacion = request.POST.get('observacionHidden', None)
    return redirect(ubicacion)


def ubicacion(request):
    
    global tipo_visitante_id
    lista_lugares = Lugar.objects.all()
    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'lista_lugares': lista_lugares,
        'tipo_visitante_id': tipo_visitante_id,
    }

    return render(request, 'ubicacion.html', context)
    

def procesar_post_ubicacion(request):
    global lugar_ID
    if request.method == 'POST':
        lugar_ID = request.POST.get('idLugar', None)
    return redirect(guardar)


def guardar(request):    
    global observacion, lugar_ID, hora_actual, trabajador_ID, fecha_actual
    
    tipo_visitante = Tipo.objects.get(tipo_ID=tipo_visitante_id)
    tipo_lugar = Lugar.objects.get(lugar_ID=lugar_ID)
    tipo_colaborador_nombre= trabajador_ID
    tipo_colaborador_apellido= ""
    tipo_colaborador_apellido2=""
    tipo_empresa=empresa_ID

    if observacion:
        observacion=observacion
    else:
        observacion=""

    if trabajador_ID:
        tipo_colaborador_id = Trabajador.objects.get(trabajador_ID=trabajador_ID)
        tipo_colaborador_nombre= tipo_colaborador_id.trabajador_nombre
        tipo_colaborador_apellido= tipo_colaborador_id.trabajador_apellido
        tipo_colaborador_apellido2= tipo_colaborador_id.trabajador_apellido2
    else:
        tipo_colaborador_nombre=""

    if empresa_ID:
        tipo_empresa_id = Empresa.objects.get(empresa_ID=empresa_ID)
        tipo_empresa= tipo_empresa_id.empresa_nombre
    else:
        tipo_empresa=""


    context = {
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'RUT': RUT,
        'observacion': observacion,
        'fecha_actual': fecha_actual,
        'fecha_salida': fecha_salida,
        'hora_actual': hora_actual,
        'tipo_visitante_id': tipo_visitante.tipo_nombre,
        'lugar_ID': tipo_lugar.lugar_nombre,
        'empresa_ID': tipo_empresa,
        'trabajador_nombre': tipo_colaborador_nombre,
        'trabajador_apellido': tipo_colaborador_apellido,
        'trabajador_apellido2': tipo_colaborador_apellido2,
    }

    return render(request, 'guardar.html',context)


def guardar_visita(request):
    global RUT, nombre, apellido_paterno, apellido_materno, fecha_actual, tipo_visitante_id, observacion, lugar_ID, fecha_salida, empresa_ID, trabajador_ID
    
    nueva_visita = Visita(
        visita_rut=RUT, 
        visita_nombre=nombre, 
        visita_apellido_1=apellido_paterno, 
        visita_apellido_2=apellido_materno, 
        visita_fecha_llegada=fecha_actual, 
        tipo_ID_id=tipo_visitante_id, 
        visita_observacion=observacion,
        lugar_ID_id=lugar_ID, 
        visita_fecha_salida=fecha_salida, 
        empresa_ID_id=empresa_ID, 
        trabajador_ID_id=trabajador_ID, 
        visita_activo=True, 
        )
    nueva_visita.save()
    return redirect(index)


def formulario_filtros(request):
    lista_trabajadores = Trabajador.objects.all()
    lista_lugares = Lugar.objects.all()
    lista_tipo = Tipo.objects.all()
    lista_empresa = Empresa.objects.all()
    lista_rut = Visita.objects.values('visita_rut').annotate(count=Count('visita_rut')).order_by('visita_rut')
    unique_rut_values = [item['visita_rut'] for item in lista_rut] #ruts unicos

    context = {
        'lista_trabajadores': lista_trabajadores,
        'lista_lugares': lista_lugares,
        'lista_tipo': lista_tipo,
        'lista_empresa': lista_empresa,
        'lista_rut': unique_rut_values,
    }
    return render(request, 'formulario_filtros.html', context)


def reporte_personalizado(request):
    global visitas_filtradas

    if request.method == 'POST':
        rut = request.POST.get('RUT', None)
        tipo = request.POST.get('tipo', None)
        empresa = request.POST.get('empresa', None)
        trabajador_id = request.POST.get('idTrabajador', None)  # Se obtiene el ID del trabajador
        lugar = request.POST.get('Ubicacion', None)
        fecha_inicio = request.POST.get('fecha_inicio', None)
        fecha_fin = request.POST.get('fecha_fin', None)
        visitas_activas = request.POST.get('visitas_activas', None)  # nuevo campo
        visitas_filtradas = Visita.objects.all()

        if tipo:
            visitas_filtradas = visitas_filtradas.filter(tipo_ID__tipo_nombre=tipo)
        if empresa:
            visitas_filtradas = visitas_filtradas.filter(empresa_ID__empresa_nombre=empresa)
        if trabajador_id:  # Filtrar por el ID del trabajador
            visitas_filtradas = visitas_filtradas.filter(trabajador_ID=trabajador_id)
        if lugar:
            visitas_filtradas = visitas_filtradas.filter(lugar_ID__lugar_nombre=lugar)
        if fecha_fin:
            fecha_fin_datetime = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d') + datetime.timedelta(days=1)
            fecha_fin = fecha_fin_datetime.strftime('%Y-%m-%d')
            visitas_filtradas = visitas_filtradas.filter(visita_fecha_llegada__lt=fecha_fin)
        if fecha_inicio:
            visitas_filtradas = visitas_filtradas.filter(visita_fecha_llegada__gte=fecha_inicio)
        if rut:
            visitas_filtradas = visitas_filtradas.filter(visita_rut=rut)

        # Agregar lógica para filtrar visitas activas
        if visitas_activas:
            visitas_filtradas = visitas_filtradas.filter(visita_activo=True)

        return redirect(reporte)  
    else:
        return render(request, 'formulario_filtros.html')


def reporte(request):
    global visitas_filtradas
    visitas = visitas_filtradas

    paginator = Paginator(visitas, 15)  # Mostrar 5 visitas por página

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'visitas_filtradas': visitas_filtradas,
        'page_obj': page_obj,
    }
    return render(request, 'reporte_personalizado.html', context)


def reporte_total(request):
    global visitas_filtradas
    visitas = Visita.objects.all()
    visitas_filtradas=visitas

    paginator = Paginator(visitas, 15)  # Mostrar 5 visitas por página

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'visitas_filtradas': visitas,
        'page_obj': page_obj,
    }
    return render(request, 'reporte_total.html', context)


def reporte_visitas_activas(request):
    global nombre,apellido_paterno,apellido_materno,RUT,hora_actual,fecha_actual

    global visitas_filtradas
    visitas_filtradas = Visita.objects.all()

    # Agregar lógica para filtrar visitas activas
    
    visitas_filtradas = visitas_filtradas.filter(visita_activo=True)

    visitas = visitas_filtradas
    paginator = Paginator(visitas, 15)  # Mostrar 5 visitas por página

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'visitas_filtradas': visitas_filtradas,
        'page_obj': page_obj,
    }
    return render(request, 'reporte_visitas_activos.html', context)


def exportar_a_excel(request):
    global visitas_filtradas
    # Crear un libro de Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe"

    # Agregar encabezados
    header_row = ['Nombre', 'Apellido', 'RUT', 'Fecha llegada', 'Hora llegada', 'Hora Salida','Duracion', 'Empresa', 'Lugar', 'Tipo de visitante', 'Colaborador', 'RUT Colaborador','Observacion']
    ws.append(header_row)

    # Agregar datos de visitas
    for visita in visitas_filtradas:

        tipo_empresa = visita.empresa_ID_id
        tipo_lugar = visita.lugar_ID_id
        tipo_visitante = visita.tipo_ID_id
        tipo_trabajador = visita.trabajador_ID_id
        trabajador_rut=""
        diferencia_tiempo=""


        if visita.empresa_ID_id :
            tipo_empresa = Empresa.objects.get(empresa_ID=visita.empresa_ID_id).empresa_nombre
        if visita.lugar_ID_id:
            tipo_lugar = Lugar.objects.get(lugar_ID=visita.lugar_ID_id).lugar_nombre
        if visita.tipo_ID_id:
            tipo_visitante = Tipo.objects.get(tipo_ID=visita.tipo_ID_id).tipo_nombre
        if visita.trabajador_ID_id:
            tipo_trabajador_nombre = Trabajador.objects.get(trabajador_ID=visita.trabajador_ID_id).trabajador_nombre
            tipo_trabajador_apellido=Trabajador.objects.get(trabajador_ID=visita.trabajador_ID_id).trabajador_apellido
            tipo_trabajador_apellido2=Trabajador.objects.get(trabajador_ID=visita.trabajador_ID_id).trabajador_apellido2
            tipo_trabajador=f"{tipo_trabajador_nombre} {tipo_trabajador_apellido} {tipo_trabajador_apellido2}"
            trabajador_rut = visita.trabajador_ID_id

        if visita.visita_fecha_salida:
            diferencia_tiempo=visita.visita_fecha_salida-visita.visita_fecha_llegada
            fecha_hora_salida=visita.visita_fecha_salida.strftime('%H:%M')
            duracion_visita=diferencia_tiempo

        else:
            duracion_visita="Aun activo"
            fecha_hora_salida="Aun activo"
        

        ws.append([
            visita.visita_nombre,
            visita.visita_apellido_1,
            visita.visita_rut,
            visita.visita_fecha_llegada.strftime("%d/%m/%Y"),
            visita.visita_fecha_llegada.strftime('%H:%M'),
            fecha_hora_salida,
            duracion_visita,
            tipo_empresa,
            tipo_lugar,
            tipo_visitante,
            tipo_trabajador,
            trabajador_rut,
            visita.visita_observacion,
        ])

    # Resaltar la primera fila
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Resaltar bordes de todas las celdas
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
    # Ajustar el ancho de las columnas al contenido        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # obtiene la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length ) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=informe.xlsx'
    wb.save(response)

    return response


